from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


THIN = Side(style="thin", color="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="0F766E")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SOLID_FILL = PatternFill("solid", fgColor="DCFCE7")
PARTIAL_FILL = PatternFill("solid", fgColor="FEF3C7")
API_FILL = PatternFill("solid", fgColor="FCE7F3")

FEATURE_HEADERS = [
    "No",
    "Feature",
    "Area",
    "Status",
    "Channel",
    "Summary",
    "Main Files",
    "Depends On API",
    "Main Gaps / Notes",
]

API_HEADERS = [
    "Feature No",
    "Feature",
    "Provider / Dependency",
    "Current Gap",
    "Main Files",
    "Implementation Tasks",
    "Priority",
]

MODULE_HEADERS = [
    "Layer",
    "Folder / Module",
    "What It Does",
    "Best Entry Files",
]

FEATURE_ROWS = [
    [
        1,
        "Catat uang via chat",
        "Core Transactions",
        "Solid",
        "WhatsApp",
        "Mencatat pemasukan dan pengeluaran dari bahasa natural user.",
        "process-inbound.ts; text-handler.ts; transaction-service.ts",
        "No",
        "Bisa ditambah coverage slang ekstrem, tapi core flow sudah stabil.",
    ],
    [
        2,
        "Foto struk / screenshot dicatat",
        "OCR",
        "Solid",
        "WhatsApp",
        "Baca image lalu ekstrak nominal dan detail transaksi.",
        "image-handler.ts; OCR pipeline",
        "No",
        "Akurasi OCR masih tergantung kualitas image.",
    ],
    [
        3,
        "Edit / hapus via chat",
        "Transaction Mutation",
        "Solid",
        "WhatsApp",
        "Ubah atau hapus transaksi dengan bahasa natural dan ambiguity handling.",
        "transaction-mutation-command-service.ts; conversation-memory-service.ts",
        "No",
        "Kasus sangat ambigu sudah lebih aman karena clarification flow.",
    ],
    [
        4,
        "Auto kategori transaksi",
        "Categorization",
        "Solid",
        "WhatsApp",
        "Transaksi otomatis masuk bucket besar dan detail tag internal.",
        "category-override-service.ts; detail-tag-service.ts; merchant-normalization-service.ts",
        "No",
        "Per-user correction belum penuh, tapi auto mapping sudah kuat.",
    ],
    [
        5,
        "Laporan harian / mingguan / bulanan",
        "Reporting",
        "Solid",
        "WhatsApp / Admin",
        "Summary report, bucket detail, analytics, compare period, dan monthly PDF.",
        "report.ts; report-service.ts; monthly-report-pdf-service.ts",
        "No",
        "Monthly full-calendar report sekarang PDF, bukan chart image.",
    ],
    [
        6,
        "Pantau budget",
        "Budgeting",
        "Solid",
        "WhatsApp",
        "Set budget, monitor usage, dan warning budget.",
        "budget-service.ts; onboarding budget flow; report-service.ts",
        "No",
        "Simulasi budget advanced masih bisa diperdalam.",
    ],
    [
        7,
        "Reminder & notifikasi pintar",
        "Reminders",
        "Partial",
        "WhatsApp / Background",
        "Budget alert, cashflow warning, weekly review, monthly closing, goal reminders, snooze, cap harian.",
        "reminder-service.ts; reminder-preference-service.ts; schema ReminderEvent",
        "No",
        "Masih bisa diperdalam untuk anti-spam policy dan smarter prioritization lintas event.",
    ],
    [
        8,
        "Target keuangan",
        "Planning / Goals",
        "Solid",
        "WhatsApp / Onboarding",
        "Multi-goal onboarding dan runtime goal management sudah aktif.",
        "goal-service.ts; goal-planner-service.ts; onboarding-service.ts",
        "No",
        "Planner lebih kaya sudah ada, tinggal fine tuning strategi adaptif.",
    ],
    [
        9,
        "Progress tabungan otomatis",
        "Planning / Goals",
        "Solid",
        "WhatsApp",
        "Progress goal sekarang pakai contribution tracking per goal, ETA, streak, tracking status.",
        "goal-service.ts; formatters.ts; goal contribution schema",
        "No",
        "Belum full ledger investasi, tapi progress goal tabungan sudah usable.",
    ],
    [
        10,
        "AI Advisor",
        "Assistant",
        "Solid",
        "WhatsApp",
        "Advice dan insight berbasis data user, bukan template statis.",
        "advice-service.ts; insight-service.ts; general-chat-service.ts",
        "No",
        "Bisa diperdalam lagi untuk reasoning lintas horizon waktu yang lebih panjang.",
    ],
    [
        11,
        "Pola pengeluaran & kebiasaan",
        "Analytics",
        "Partial",
        "WhatsApp / Admin",
        "Top merchant, leak detection, recurring hints, explainable analytics, weekend vs weekday.",
        "report-service.ts; insight-service.ts",
        "No",
        "Masih bisa diperdalam untuk longitudinal root-cause analysis multi-bulan.",
    ],
    [
        12,
        "Portfolio investasi",
        "Portfolio",
        "Partial",
        "WhatsApp",
        "Catat aset, nilai aset, komposisi, konsentrasi, diversification hint, rebalance hint dasar.",
        "portfolio-command-service.ts; portfolio-valuation-service.ts",
        "No",
        "Belum penuh untuk trade ledger, realized P/L, cost basis detail, dan analytics portfolio advanced.",
    ],
    [
        13,
        "Update harga market",
        "Market Data",
        "Requires API",
        "WhatsApp",
        "Lookup harga saham, crypto, emas, dan valuasi live portfolio.",
        "market-price-service.ts; market-command-service.ts; portfolio-valuation-service.ts",
        "Yes",
        "Provider final belum di-wire penuh.",
    ],
    [
        14,
        "Berita harian finance",
        "News",
        "Requires API",
        "WhatsApp",
        "Digest berita finance harian dengan filtering yang layak produksi.",
        "finance-news-service.ts",
        "Yes",
        "Masih perlu provider news production-grade.",
    ],
    [
        15,
        "News personal ke portfolio",
        "News / Portfolio",
        "Requires API",
        "WhatsApp",
        "Filter berita berdasarkan aset user dan jelaskan dampak potensialnya.",
        "finance-news-service.ts; portfolio-command-service.ts",
        "Yes",
        "Butuh entity/symbol matching dan provider news yang lebih akurat.",
    ],
    [
        16,
        "Smart allocation",
        "Planning",
        "Partial",
        "WhatsApp",
        "Rekomendasi alokasi sisa uang, goal split, dan prioritas tabungan.",
        "smart-allocation-service.ts; goal-planner-service.ts",
        "No",
        "Masih rule-based, belum adaptive / risk-profile driven penuh.",
    ],
    [
        17,
        "Financial freedom tracker",
        "Planning",
        "Partial",
        "WhatsApp",
        "Tracker FF, gap analysis, target lean/base/conservative, dan planner dasar.",
        "financial-freedom-service.ts",
        "No",
        "Masih baseline planner, belum simulator asumsi lanjutan penuh.",
    ],
    [
        18,
        "Wealth projection",
        "Planning",
        "Partial",
        "WhatsApp",
        "Simulasi nilai masa depan, target reach, compare skenario, dan starting principal.",
        "wealth-projection-service.ts; projection-math-service.ts",
        "No",
        "Belum semua skenario kompleks bertahap dan inflasi detail.",
    ],
]

API_ROWS = [
    [
        13,
        "Update harga market",
        "Finnhub + GoldAPI + exchangerate.host",
        "Quote live belum production-grade dan symbol normalization belum final.",
        "market-price-service.ts; market-command-service.ts; portfolio-valuation-service.ts",
        "Buat provider abstraction; tambah symbol normalization; cache + stale fallback; wire quote ke market command; pakai untuk portfolio valuation live; tambah graceful no-key fallback; tambah observability latency/error.",
        "P0",
    ],
    [
        14,
        "Berita harian finance",
        "Marketaux (primary) + RSS fallback",
        "Digest masih basic dan belum pakai provider finance-grade.",
        "finance-news-service.ts; structured-text-handler.ts; global-context-router-service.ts",
        "Buat provider client; filter by locale/category; rank headline; format digest ringkas; fallback ke RSS kalau API gagal; tambah test untuk no-key dan provider down.",
        "P0",
    ],
    [
        15,
        "News personal ke portfolio",
        "Marketaux + portfolio symbol/entity mapping",
        "Belum ada personalization yang cukup akurat per aset user.",
        "finance-news-service.ts; portfolio-command-service.ts; market-price-service.ts",
        "Map asset type -> symbol/entity; score relevance; attach impact summary per asset; handle user with mixed local/global assets; tambah fallback bila symbol tidak tersedia.",
        "P0",
    ],
    [
        "Cross-cutting",
        "Env & Config",
        "API keys + feature flags",
        "Belum ada setup final untuk provider live.",
        "env.ts; deployment env; secrets management",
        "Tambah env validation untuk FINNHUB_API_KEY, GOLDAPI_API_KEY, MARKETAUX_API_TOKEN, EXCHANGERATE_API_KEY; siapkan feature flag dan fallback behavior saat key kosong.",
        "P0",
    ],
    [
        "Cross-cutting",
        "Testing & Monitoring",
        "Provider mocks + observability",
        "Belum ada contract tests penuh untuk provider live path.",
        "tests market/news; observability-service.ts",
        "Tambah mocked integration tests; test fallback chain; log provider hit rate, cache hit, latency, failure count; tampilkan di observability admin bila perlu.",
        "P1",
    ],
]

MODULE_ROWS = [
    [
        "Inbound Entry",
        "apps/api/lib/features/inbound",
        "Pintu masuk pesan user dari WA, routing onboarding, text, image, report response.",
        "process-inbound.ts; text-handler.ts; structured-text-handler.ts; report.ts; image-handler.ts",
    ],
    [
        "Assistant Layer",
        "apps/api/lib/services/assistant",
        "Intent routing global, semantic normalization, memory, ambiguity resolution, general chat safe fallback.",
        "global-context-router-service.ts; conversation-memory-service.ts; general-chat-service.ts; ai-service.ts",
    ],
    [
        "Transactions",
        "apps/api/lib/services/transactions",
        "Pencatatan transaksi, mutation, kategori, merchant normalization, detail tag.",
        "transaction-service.ts; transaction-mutation-command-service.ts; category-override-service.ts; merchant-normalization-service.ts",
    ],
    [
        "Reporting",
        "apps/api/lib/services/reporting",
        "Report summary, analytics, insight, health score, monthly PDF.",
        "report-service.ts; insight-service.ts; advice-service.ts; monthly-report-pdf-service.ts",
    ],
    [
        "Planning",
        "apps/api/lib/services/planning",
        "Goals, contributions, smart allocation, financial freedom, wealth projection.",
        "goal-service.ts; goal-planner-service.ts; smart-allocation-service.ts; financial-freedom-service.ts; wealth-projection-service.ts",
    ],
    [
        "Portfolio / Market / News",
        "apps/api/lib/services/market",
        "Assets, valuation, market lookup, finance news, portfolio analytics.",
        "portfolio-command-service.ts; portfolio-valuation-service.ts; market-price-service.ts; finance-news-service.ts",
    ],
    [
        "Reminders",
        "apps/api/lib/services/reminders",
        "Reminder engine, preference, snooze, daily cap, review digest.",
        "reminder-service.ts; reminder-preference-service.ts",
    ],
    [
        "Onboarding",
        "apps/api/lib/services/onboarding",
        "Branching onboarding flow, parsing jawaban, persist profile, initial analysis.",
        "onboarding-service.ts; onboarding-flow-service.ts; onboarding-parser-service.ts",
    ],
    [
        "Admin API",
        "apps/api/app/api/admin",
        "HTTP surface untuk admin web: users, transactions, subscriptions, health, observability.",
        "users/route.ts; transactions/route.ts; subscriptions/route.ts; health/route.ts; observability/route.ts",
    ],
    [
        "Admin Web",
        "apps/admin-web",
        "Dashboard internal untuk monitoring operasional dan audit data.",
        "app/layout.tsx; app/users/page.tsx; app/transactions/page.tsx; app/observability/page.tsx",
    ],
    [
        "WhatsApp Transport",
        "apps/bot/src",
        "Pengiriman text/document ke WhatsApp dan penerimaan callback bot.",
        "index.ts",
    ],
    [
        "Reporting Service",
        "services/reporting",
        "Service Python untuk generate artefak report seperti PDF monthly.",
        "app/main.py",
    ],
]


def style_sheet(ws, status_col=None, api_col=None, priority_col=None):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        if status_col is not None:
            status_value = row[status_col - 1].value
            if status_value == "Solid":
                row[status_col - 1].fill = SOLID_FILL
            elif status_value == "Partial":
                row[status_col - 1].fill = PARTIAL_FILL
            elif status_value == "Requires API":
                row[status_col - 1].fill = API_FILL

        if api_col is not None and row[api_col - 1].value == "Yes":
            row[api_col - 1].fill = API_FILL

        if priority_col is not None and row[priority_col - 1].value == "P0":
            row[priority_col - 1].fill = PARTIAL_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value


base = Path(__file__).resolve().parent.parent
out_dir = base / "artifacts"
out_dir.mkdir(exist_ok=True)
out_file = out_dir / "ai-finance-assistant-feature-inventory.xlsx"

wb = Workbook()
wb.remove(wb.active)

summary = wb.create_sheet("00_Summary")
summary["A1"] = "AI Finance Assistant - Feature Inventory"
summary["A1"].font = Font(size=16, bold=True)
summary["A3"] = "Generated At"
summary["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
summary["A4"] = "Purpose"
summary["B4"] = "Feature inventory, current status, API backlog, and code module map for developer handoff."
summary["A5"] = "Status Legend"
summary["B5"] = "Solid = ready core flow; Partial = usable but still needs upgrade; Requires API = blocked on live provider integration."
summary.append([])
summary.append(["Sheet", "Purpose", "Row Count"])
for cell in summary[7]:
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
summary_rows = [
    ("01_Feature_Matrix", "Complete feature list with status, channels, files, and notes", len(FEATURE_ROWS)),
    ("02_API_Backlog", "Tasks for API-dependent partial features", len(API_ROWS)),
    ("03_Module_Map", "Developer-oriented map of the code structure", len(MODULE_ROWS)),
]
for item in summary_rows:
    summary.append(list(item))
set_widths(summary, {"A": 24, "B": 88, "C": 16})
summary.freeze_panes = "A7"

feature_sheet = wb.create_sheet("01_Feature_Matrix")
feature_sheet.append(FEATURE_HEADERS)
for row in FEATURE_ROWS:
    feature_sheet.append(row)
style_sheet(feature_sheet, status_col=4, api_col=8)
set_widths(
    feature_sheet,
    {
        "A": 6,
        "B": 34,
        "C": 20,
        "D": 16,
        "E": 18,
        "F": 44,
        "G": 56,
        "H": 14,
        "I": 44,
    },
)

api_sheet = wb.create_sheet("02_API_Backlog")
api_sheet.append(API_HEADERS)
for row in API_ROWS:
    api_sheet.append(row)
style_sheet(api_sheet, priority_col=7)
set_widths(
    api_sheet,
    {
        "A": 14,
        "B": 28,
        "C": 28,
        "D": 42,
        "E": 56,
        "F": 64,
        "G": 10,
    },
)

module_sheet = wb.create_sheet("03_Module_Map")
module_sheet.append(MODULE_HEADERS)
for row in MODULE_ROWS:
    module_sheet.append(row)
style_sheet(module_sheet)
set_widths(module_sheet, {"A": 20, "B": 32, "C": 48, "D": 56})

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = True

wb.save(out_file)
loaded = load_workbook(out_file)
print(out_file)
print(",".join(loaded.sheetnames))
for name in loaded.sheetnames:
    ws = loaded[name]
    print(f"{name}:{ws.max_row - 1}")
