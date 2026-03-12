from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

COMMON_HEADERS = [
    "TC_ID","Journey","Feature","Subfeature","Support_Status","Priority",
    "Preconditions","User_Input_or_Action","Expected_Route_or_Module",
    "Expected_Response_or_Output","Expected_Data_or_Side_Effect","Notes"
]
QUERY_HEADERS = [
    "Feature","Intent","Canonical_Command_or_Query","Variant_Query",
    "Expected_Module","Expected_Output_Type","Requires_API","Notes"
]

def parse_tsv(text, expected_cols):
    rows = []
    for raw in text.strip().splitlines():
        line = raw.strip()
        if not line or line.startswith('#'):
            continue
        parts = [part.strip() for part in line.split('\t')]
        if len(parts) < expected_cols:
            parts.extend([''] * (expected_cols - len(parts)))
        rows.append(parts[:expected_cols])
    return rows

THIN = Side(style='thin', color='D9D9D9')
HEADER_FILL = PatternFill('solid', fgColor='0F766E')
HEADER_FONT = Font(color='FFFFFF', bold=True)
SOLID_FILL = PatternFill('solid', fgColor='DCFCE7')
PARTIAL_FILL = PatternFill('solid', fgColor='FEF3C7')
API_FILL = PatternFill('solid', fgColor='FCE7F3')
P0_FILL = PatternFill('solid', fgColor='FEF3C7')


def style_sheet(ws, common=True, api_col=None):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
        if common:
            if row[4].value == 'Solid':
                row[4].fill = SOLID_FILL
            elif row[4].value == 'Partial':
                row[4].fill = PARTIAL_FILL
            elif row[4].value == 'Requires API':
                row[4].fill = API_FILL
            if row[5].value == 'P0':
                row[5].fill = P0_FILL
        elif api_col is not None and row[api_col - 1].value == 'Yes':
            row[api_col - 1].fill = API_FILL
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def set_widths(ws, widths):
    for key, value in widths.items():
        ws.column_dimensions[key].width = value
FIRST_USER = parse_tsv("""
FTU-001	First visit	Welcome	Sapaan awal	Solid	P0	User baru	Hi	Onboarding welcome	Bot kirim welcome onboarding natural	onboarding_status in_progress	Tidak nyasar ke transaksi
FTU-002	First visit	Welcome	Start natural	Solid	P0	Welcome sudah terkirim	gaskeun boss	Onboarding semantic start	Bot lanjut ke tujuan utama	onboarding_sessions upsert	Jawaban bebas harus dipahami
FTU-003	Onboarding	Primary goal	Semua di atas	Solid	P0	Step goal aktif	semua di atas	Onboarding goal	Bot simpan primary_goal lalu lanjut employment	users.primary_goal terisi	
FTU-004	Onboarding	Employment	Mixed role	Solid	P0	Step employment aktif	saya karyawan sambil usaha kecil	Onboarding employment	Bot pahami mixed role	users.employment_type & stability	
FTU-005	Onboarding	Income	Salary + passive	Solid	P0	Employee path	gaji 8 juta, tgl 25, passive 3 juta	Onboarding income	Bot pecah jawaban sesuai step	financial_profile & salary_date terisi	
FTU-006	Onboarding	Budget plan	Flexible manual	Solid	P0	Step budget manual	makan 1jt, transport 200rb, tagihan 350rb, istri 1jt	Onboarding expense parser	Bot terima format bebas dan map label asing ke others	expense plan items tercipta	Regression parser budget
FTU-007	Onboarding	Goals	Multiple goals	Solid	P0	Step target aktif	beli rumah 500 juta, lalu liburan 15 juta	Onboarding goals	Bot buat beberapa goal	financial_goals multi rows	
FTU-008	Onboarding	Assets	Gold + savings	Partial	P1	Step aset aktif	emas antam 10 gram dan tabungan 20 juta	Onboarding assets	Bot simpan dua aset	assets rows tercipta	Valuasi live perlu API
FTU-009	Onboarding	Summary	Initial analysis	Solid	P0	Onboarding selesai	ga ada lagi	Onboarding completion	Bot kirim summary awal lengkap	analysis_ready true	Summary harus jelaskan grouping
FTU-010	First usage	Transaction	First expense	Solid	P0	Onboarding complete	beli kopi 25 ribu	Record transaction	Bot catat expense	transaction EXPENSE tersimpan	
FTU-011	First usage	Transaction	First income	Solid	P0	Onboarding complete	gaji masuk 8 juta	Record transaction	Bot catat income	transaction INCOME tersimpan	
FTU-012	First usage	Report	Monthly PDF	Solid	P0	Ada transaksi & reporting service aktif	laporan bulan ini	Monthly report	Bot kirim text summary + PDF	No DB change	Gambar tidak dipakai
""", len(COMMON_HEADERS))

ONBOARDING = parse_tsv("""
ONB-001	Onboarding	Welcome	No option list	Solid	P0	User baru	Halo	Onboarding welcome	Single CTA tanpa Pilihan:	No data mutation	
ONB-002	Onboarding	Welcome	Natural affirmative	Solid	P0	Welcome terkirim	oke saya siap	Onboarding start	Lanjut ke tujuan utama	onboarding_sessions upsert	
ONB-003	Onboarding	Employment	Mahasiswa	Solid	P1	Step employment aktif	mahasiswa	Onboarding employment	Tanya estimasi income	users.employment_type terisi	
ONB-004	Onboarding	Employment	Freelance	Solid	P1	Step employment aktif	freelance	Onboarding employment	Tanya estimasi income + passive	users.employment_type terisi	
ONB-005	Onboarding	Employment	Karyawan	Solid	P0	Step employment aktif	karyawan	Onboarding employment	Tanya active income + tanggal gajian	users.salary_date pending	
ONB-006	Onboarding	Income	Passive yes	Solid	P0	Step passive aktif	ada	Onboarding boolean	Lanjut minta nominal passive	users.has_passive_income true	
ONB-007	Onboarding	Income	Passive no natural	Solid	P0	Step passive aktif	ga ada	Onboarding boolean	Lanjut step berikut, tidak loop	users.has_passive_income false	Regression ga ada
ONB-008	Onboarding	Income	Nominal shorthand	Solid	P0	Step nominal aktif	3jt	Onboarding amount parse	Bot pahami 3000000	passive_income_monthly numeric	
ONB-009	Onboarding	Budget mode	Guided	Solid	P0	Step budget mode aktif	belum punya tapi mau dibantu	Onboarding budget mode	Bot tanya kategori satu per satu	users.budget_mode guided_plan	
ONB-010	Onboarding	Budget mode	Auto	Solid	P0	Step budget mode aktif	belum punya dan mau dianalisis otomatis	Onboarding budget mode	Bot jelaskan belajar dari transaksi	budget_mode auto_from_transactions	
ONB-011	Onboarding	Expense allocation	Flexible multiline	Solid	P0	Budget manual aktif	makan 1jt\ntransport 200rb\ntagihan 350rb\nistri 1jt	Onboarding expense parser	Bot terima dan map istri ke others	expense_plan_items tercipta	Tidak boleh stuck
ONB-012	Onboarding	Goals	Emergency fund with expense	Solid	P0	Expense tersedia	dana darurat	Onboarding goal creation	Target dana darurat dihitung otomatis	goal target_amount auto	
ONB-013	Onboarding	Goals	Emergency fund pending	Solid	P0	Expense belum ada	dana darurat	Pending calculation flow	Bot tawarkan bantu hitung / sudah punya data / lewati	goal pending_calculation bila dilewati	
ONB-014	Onboarding	Goals	Financial freedom age	Solid	P1	Expense tersedia	financial freedom usia 45	Onboarding FF goal	Bot hitung target dan simpan usia	FF goal + target age tersimpan	
ONB-015	Onboarding	Goals	No more goals	Solid	P0	Step ada lagi? aktif	ga ada lagi	Onboarding boolean	Bot pindah ke aset	Session step selesai	Regression looping target
ONB-016	Onboarding	Assets	No assets	Solid	P1	Step aset aktif	belum punya	Onboarding assets	Bot tandai belum punya aset	users.has_assets false	
ONB-017	Onboarding	Resume	Resume last step	Solid	P0	User keluar di tengah onboarding	lanjut	Onboarding resume	Bot lanjut dari current step	Session dipakai resume	
ONB-018	Onboarding	Analysis	Final summary	Solid	P0	Onboarding selesai	selesai	Onboarding summary	Summary rinci income, expense, goals, aset, grouping	analysis context siap	
""", len(COMMON_HEADERS))

TRANSACTIONS = parse_tsv("""
TRX-001	Daily use	Transaction	Expense simple	Solid	P0	User aktif	beli kopi 25 ribu	Record transaction	Catat expense Food & Drink	transaction EXPENSE	
TRX-002	Daily use	Transaction	Income simple	Solid	P0	User aktif	gaji masuk 5 juta	Record transaction	Catat income	transaction INCOME	
TRX-003	Daily use	Transaction	Natural phrasing	Solid	P0	User aktif	tadi makan siang abis 45rb	Semantic -> transaction	Tetap tercatat sebagai expense	transaction tersimpan	
TRX-004	Daily use	Transaction	Forced category	Solid	P1	User aktif	bayar parkir 5rb kategori transport	Transaction + override	Masuk Transport	transaction category updated	
TRX-005	Daily use	OCR	Receipt image	Solid	P1	OCR service aktif	Kirim foto struk belanja	OCR -> transaction	Bot baca nominal dan detail	transaction OCR	
TRX-006	Daily use	Mutation	Edit recent	Solid	P0	Ada transaksi kopi	ubah kopi tadi jadi 20 ribu	Transaction mutation	Bot update transaksi benar	transaction updated	
TRX-007	Daily use	Mutation	Delete by label	Solid	P0	Ada transaksi listrik	hapus transaksi listrik	Transaction mutation	Bot hapus transaksi cocok	transaction deleted	
TRX-008	Daily use	Mutation	Delete latest	Solid	P0	Ada transaksi terakhir	hapus yang barusan	Transaction mutation + memory	Hapus transaksi terbaru relevan	transaction deleted	
TRX-009	Daily use	Mutation	Ambiguous delete	Solid	P0	Ada beberapa Spotify	hapus transaksi spotify	Ambiguity resolver	Bot minta pilih kandidat	Tidak ada delete dulu	
TRX-010	Daily use	Mutation	Clarification follow-up	Solid	P0	Bot baru kirim kandidat	yang kedua	Clarification resolver	Eksekusi kandidat nomor 2	transaction target terhapus	
TRX-011	Daily use	Category	Merchant normalization	Solid	P1	Ada histori Spotify	spotify premium 54.990	Merchant normalization	Normalisasi ke merchant Spotify	merchant normalized	
TRX-012	Daily use	Category	Detail tag stored	Solid	P1	User aktif	bayar spotify 50rb	Detail tagging	Bucket tetap Entertainment, detail Spotify tersimpan	detailTag tersimpan	
TRX-013	Daily use	Safety	Out of scope	Solid	P0	User aktif	cuaca hari ini gimana	General chat safe	Bot tidak halu dan jelaskan scope	No data change	
TRX-014	Daily use	Safety	Ambiguous utterance	Solid	P0	Tanpa konteks	 yang tadi aja	Clarification fallback	Bot minta klarifikasi	No data change	
TRX-015	Daily use	OCR	Blur fallback	Solid	P2	OCR gagal	Kirim gambar blur	OCR safe fallback	Bot minta kirim ulang / tulis manual	Tidak ada transaksi	
""", len(COMMON_HEADERS))
REPORTS = parse_tsv("""
RPT-001	Reporting	Report	Daily report	Solid	P0	Ada transaksi hari ini	laporan hari ini	General report	Bot kirim ringkasan harian text-only	No data change	
RPT-002	Reporting	Report	Weekly report	Solid	P0	Ada transaksi minggu ini	summary minggu ini	General report	Bot kirim ringkasan mingguan text-only	No data change	
RPT-003	Reporting	Report	Monthly report PDF	Solid	P0	Ada transaksi bulanan + reporting service aktif	laporan bulan ini	General report monthly	Bot kirim summary + PDF	No data change	
RPT-004	Reporting	Report	Explicit month	Solid	P1	Ada transaksi Januari 2026	laporan januari 2026	Date range report	Bot kirim summary + PDF bila full month	No data change	
RPT-005	Reporting	Report	Partial range	Solid	P1	Ada transaksi 1-15 Maret	summary 1-15 maret 2026	Partial range report	Bot kirim text-only	No data change	
RPT-006	Reporting	Report	Comparison	Solid	P1	Ada 2 periode	laporan 3 bulan terakhir vs 3 bulan sebelumnya	Comparison report	Bot kirim text compare tanpa PDF	No data change	
RPT-007	Reporting	Bucket detail	List detail	Solid	P0	Ada entertainment	 detail entertainment bulan ini apa saja	Category detail LIST	Bot kirim daftar transaksi entertainment	No data change	
RPT-008	Reporting	Bucket detail	Top transaction	Solid	P1	Ada entertainment	entertainment terbesar bulan ini apa	Category detail TOP	Bot kirim transaksi terbesar	No data change	
RPT-009	Reporting	Bucket detail	Total filtered detail	Solid	P1	Ada Spotify	spotify bulan ini total berapa	Category detail TOTAL	Bot hitung total Spotify	No data change	
RPT-010	Reporting	Bucket detail	Count filtered detail	Solid	P1	Ada Spotify	berapa transaksi spotify bulan ini	Category detail COUNT	Bot hitung jumlah transaksi	No data change	
RPT-011	Reporting	Bucket detail	Average weekly	Solid	P1	Ada bills	rata-rata spending bills per minggu	Category detail AVERAGE_WEEKLY	Bot kirim rata-rata weekly	No data change	
RPT-012	Reporting	Bucket detail	Share of bucket	Solid	P1	Ada Spotify + Entertainment	spotify kontribusinya berapa persen dari entertainment bulan ini	Category detail SHARE_OF_BUCKET	Bot kirim kontribusi persen	No data change	
RPT-013	Reporting	Analytics	Top category increase	Solid	P0	Ada current vs previous	kategori mana yang paling naik dibanding bulan lalu	General analytics	Bot kirim kategori naik tertinggi	No data change	
RPT-014	Reporting	Analytics	Top merchant delta	Solid	P0	Ada current vs previous	merchant apa yang paling ngedorong kenaikan spending	General analytics	Bot kirim merchant pendorong kenaikan	No data change	
RPT-015	Reporting	Analytics	Recurring expense	Solid	P1	Ada pola recurring	top recurring expense bulan ini	General analytics	Bot kirim recurring teratas	No data change	
RPT-016	Reporting	Analytics	New merchants	Solid	P1	Ada transaksi dua periode	merchant baru bulan ini apa aja	General analytics	Bot kirim merchant/detail baru	No data change	
RPT-017	Reporting	Analytics	Weekend vs weekday	Solid	P1	Ada histori lintas hari	weekend lebih boros gak	General analytics	Bot bandingkan weekend vs weekday	No data change	
RPT-018	Reporting	Analytics	Habit leaks	Partial	P1	Ada histori cukup	kebiasaan bocor halus aku apa	General analytics	Bot identifikasi leak pattern utama	No data change	Masih bisa diperdalam
RPT-019	Reporting	Cashflow	Safe until payday	Solid	P0	Ada income/expense + salary date	aman sampai gajian gak	Cashflow forecast	Bot jawab safety sampai payday	No data change	
RPT-020	Reporting	Cashflow	Month-end remaining	Solid	P0	Ada histori bulanan	ujung bulan kira-kira sisa uang berapa	Cashflow forecast	Bot jawab estimasi sisa akhir bulan	No data change	
RPT-021	Reporting	Cashflow	Scenario expense	Solid	P1	Ada cashflow data	kalau bayar cicilan 1 juta besok masih aman gak	Cashflow scenario	Bot hitung impact skenario expense	No data change	
RPT-022	Reporting	Health	Health score	Solid	P1	Ada data bulan berjalan	skor keuangan bulan ini	Financial health score	Bot kirim health score + komponen	No data change	
RPT-023	Reporting	Health	Monthly closing	Solid	P1	Ada data bulan target	closing januari 2026	Financial health closing	Bot kirim closing bulan target	No data change	
RPT-024	Reporting	Date parser	Quarter range	Solid	P1	Ada data Q1 2026	laporan q1 2026	Date parser	Bot pahami Q1 2026	No data change	
""", len(COMMON_HEADERS))

PLANNING = parse_tsv("""
PLN-001	Planning	Budget	Set budget natural	Solid	P0	User aktif	budget makan 2 juta per bulan	Budget set	Bot simpan budget makan	budget upsert	
PLN-002	Planning	Goal	Set simple goal	Solid	P0	User aktif	mau nabung 50 juta	Goal set	Bot buat goal generic	goal target created	
PLN-003	Planning	Goal	Set house goal	Solid	P0	User aktif	target rumah 800 juta	Goal set typed	Bot buat goal rumah	goal row HOUSE	
PLN-004	Planning	Goal	Goal status generic	Solid	P0	Ada goal aktif	status tabungan aku gimana	Goal status	Bot tampilkan progress goal aktif	No data change	
PLN-005	Planning	Goal	Goal status by name	Solid	P0	Ada goal rumah	status goal rumah gimana	Goal status filtered	Bot tampilkan progress goal rumah	No data change	
PLN-006	Planning	Goal	Contribution	Solid	P0	Ada goal rumah	setor 500rb ke rumah	Goal contribution	Bot tambah kontribusi ke rumah	goal contribution row	
PLN-007	Planning	Goal planner	Focus	Solid	P1	Ada beberapa goal	kalau fokus rumah dulu gimana	Goal planner FOCUS	Bot jelaskan dampak fokus goal	No data change	
PLN-008	Planning	Goal planner	Focus duration	Solid	P1	Ada beberapa goal	kalau fokus rumah 6 bulan dulu gimana	Goal planner FOCUS_DURATION	Bot simulasi fokus 6 bulan	No data change	
PLN-009	Planning	Goal planner	Split recommendation	Solid	P1	Ada beberapa goal	tabungan bulan ini paling baik dibagi ke target apa	Goal planner SPLIT	Bot rekomendasikan alokasi per goal	No data change	
PLN-010	Planning	Goal planner	Custom split ratio	Solid	P1	Ada beberapa goal	kalau tabungan dibagi 60:40 hasilnya gimana	Goal planner SPLIT_RATIO	Bot simulasikan split 60:40	No data change	
PLN-011	Planning	Goal planner	Priority	Solid	P1	Ada beberapa goal	target mana yang paling realistis dulu	Goal planner PRIORITY	Bot beri ranking prioritas	No data change	
PLN-012	Planning	Advisor	Smart allocation	Partial	P1	Ada profile, expense, goals	sisa uang bulan ini sebaiknya kemana	Smart allocation	Bot beri rekomendasi alokasi	No data change	Masih rule-based
PLN-013	Planning	Financial freedom	Activate/track	Partial	P1	Ada expense dan income	aktifkan financial freedom	Financial freedom	Bot aktifkan tracker FF	Setting / goal FF updated	Masih baseline
PLN-014	Planning	Financial freedom	Status	Partial	P1	FF aktif	financial freedom aku aman gak	Financial freedom	Bot jawab status dan gap	No data change	Masih baseline
PLN-015	Planning	Wealth projection	Monthly invest	Partial	P1	User aktif	kalau invest 3 juta per bulan 10 tahun hasilnya berapa	Wealth projection	Bot kirim skenario proyeksi	No data change	Belum semua skenario kompleks
PLN-016	Planning	Wealth projection	Target reach	Partial	P1	User aktif	kalau target 1 miliar kapan tercapai	Wealth projection	Bot hitung estimasi target reach	No data change	
""", len(COMMON_HEADERS))

PORTFOLIO = parse_tsv("""
PRT-001	Portfolio	Assets	Add gold	Partial	P1	User aktif	tambah emas 10 gram harga 1200000 per gram	Portfolio command	Bot tambah aset gold	asset row gold	Perlu API untuk live value
PRT-002	Portfolio	Assets	Add stock	Partial	P1	User aktif	tambah saham BBCA 10 lot harga 9000	Portfolio command	Bot tambah aset stock	asset row stock	Market live belum final
PRT-003	Portfolio	Assets	Add crypto	Partial	P1	User aktif	tambah BTC 0.02 harga 1000000000	Portfolio command	Bot tambah aset crypto	asset row crypto	
PRT-004	Portfolio	Summary	Portfolio snapshot	Partial	P0	Ada aset	portfolio aku gimana	Portfolio command	Bot tampilkan total aset, komposisi, top holding	No data change	
PRT-005	Portfolio	Summary	Asset value	Partial	P0	Ada aset	nilai aset aku berapa	Portfolio command	Bot tampilkan total nilai aset	No data change	
PRT-006	Portfolio	Analytics	Dominant asset	Partial	P1	Ada aset beragam	aset paling dominan apa	Portfolio analytics	Bot tampilkan aset dominan + persen	No data change	
PRT-007	Portfolio	Analytics	Risk concentration	Partial	P1	Ada aset beragam	portfolio terlalu numpuk gak	Portfolio analytics	Bot jelaskan konsentrasi aset	No data change	
PRT-008	Portfolio	Analytics	Rebalance hint	Partial	P1	Ada aset beragam	perlu rebalance gak	Portfolio analytics	Bot beri hint rebalance	No data change	Tanpa realized P/L
PRT-009	Portfolio	Analytics	Diversification	Partial	P1	Ada aset beragam	diversifikasi portfolio aku gimana	Portfolio analytics	Bot beri skor diversifikasi	No data change	
PRT-010	Market	Price query	Stock price	Requires API	P1	Provider market aktif	harga BBCA hari ini	Market command	Bot kirim harga saham	No data change	Tergantung provider
PRT-011	Market	Fallback	Provider unavailable	Partial	P1	Provider market mati	harga BBCA hari ini	Market fallback	Bot jawab market data belum tersedia	No data change	Graceful failure
PRT-012	News	Digest	Finance daily digest	Requires API	P1	Provider news aktif	berita finance pagi ini	Finance news command	Bot kirim digest finance	No data change	Provider final belum diset
PRT-013	News	Personalized	Portfolio news	Requires API	P1	Provider news aktif + user punya aset	ada news penting tentang aset aku gak?	Finance news personalized	Bot filter news sesuai aset	No data change	
""", len(COMMON_HEADERS))

REMINDERS = parse_tsv("""
RMD-001	Reminder	Budget alert	Near limit	Partial	P1	Budget > 80% terpakai	scheduler run	Reminder engine	Bot kirim warning budget hampir habis	Reminder event tercatat	
RMD-002	Reminder	Goal alert	Off-track	Partial	P1	Goal OFF_TRACK	scheduler run	Reminder engine	Bot kirim reminder goal off-track	Reminder event tercatat	
RMD-003	Reminder	Cashflow warning	Low buffer	Partial	P1	Cashflow tipis sebelum payday	scheduler run	Reminder engine	Bot kirim warning cashflow buffer tipis	Reminder event tercatat	
RMD-004	Reminder	Weekly review	Auto digest	Solid	P2	Preference enabled	scheduler run	Weekly review	Bot kirim weekly review otomatis	Reminder event tercatat	
RMD-005	Reminder	Monthly closing	Auto closing	Solid	P2	Preference enabled + awal bulan	scheduler run	Monthly closing	Bot kirim closing bulan lalu	Reminder event tercatat	
RMD-006	Reminder	Status	Preference status	Solid	P1	User aktif	status reminder	Reminder preference	Bot tampilkan setting reminder	No data change	
RMD-007	Reminder	Disable budget	Preference update	Solid	P1	User aktif	matikan reminder budget	Reminder preference	Bot matikan budget reminder	Preference updated	
RMD-008	Reminder	Daily cap	Preference update	Solid	P1	User aktif	batasi reminder 2 per hari	Reminder preference	Bot simpan daily cap	Preference updated	
RMD-009	Reminder	Pause temporary	Preference update	Solid	P1	User aktif	pause reminder 12 jam	Reminder preference	Bot snooze sementara	Preference updated	
RMD-010	General chat	Capability	What can you do	Solid	P1	User aktif	kamu bisa apa?	General chat	Bot jelaskan kemampuan utama	No data change	
RMD-011	General chat	Safe fallback	Out of scope	Solid	P0	User aktif	tolong kerjain PR fisika	General chat safe	Bot jujur soal scope finance	No data change	
RMD-012	Conversation memory	Follow-up	Select from context	Solid	P1	Bot baru kirim daftar kandidat	yang paling tinggi aja	Conversation memory rewrite	Bot pakai konteks sebelumnya	No data change	
""", len(COMMON_HEADERS))
QUERY_ROWS = parse_tsv("""
Onboarding	Start onboarding	Mulai onboarding	oke saya siap	Onboarding	Question flow	No	
Onboarding	Start onboarding	Mulai onboarding	gaskeun boss	Onboarding	Question flow	No	
Onboarding	Start onboarding	Mulai onboarding	ayo mulai	Onboarding	Question flow	No	
Onboarding	Budget mode guided	Belum punya budget tapi mau dibantu	belum punya tapi mau dibantu	Onboarding	Question flow	No	
Onboarding	Budget mode guided	Belum punya budget tapi mau dibantu	tolong buatin budget	Onboarding	Question flow	No	
Transaction	Record expense	Catat pengeluaran	beli kopi 25 ribu	Transaction	Confirmation	No	
Transaction	Record expense	Catat pengeluaran	tadi makan siang abis 45rb	Transaction	Confirmation	No	
Transaction	Record income	Catat pemasukan	gaji masuk 5 juta	Transaction	Confirmation	No	
Transaction	Record income	Catat pemasukan	dapet fee 2 juta	Transaction	Confirmation	No	
Transaction	Edit	Ubah transaksi	ubah kopi tadi jadi 20 ribu	Transaction Mutation	Mutation result	No	
Transaction	Delete	Hapus transaksi	hapus transaksi listrik	Transaction Mutation	Mutation result	No	
Transaction	Delete	Hapus transaksi	hapus spotify yang 50rb	Transaction Mutation	Clarification if ambiguous	No	
Budget	Set budget	Atur budget kategori	budget makan 2 juta per bulan	Budget	Confirmation	No	
Report	Daily report	Laporan harian	laporan hari ini	Report	Text	No	
Report	Weekly report	Laporan mingguan	summary minggu ini	Report	Text	No	
Report	Monthly report	Laporan bulanan	laporan bulan ini	Report	Text + PDF	No	
Report	Monthly report	Laporan bulanan	laporan januari 2026	Report	Text + PDF	No	
Report	Range report	Laporan range	summary 1-15 maret 2026	Report	Text	No	
Report	Comparison	Bandingkan periode	laporan 3 bulan terakhir vs 3 bulan sebelumnya	Report	Text	No	
Analytics	Bucket detail	Rincian bucket	detail entertainment bulan ini apa saja	Category Detail Report	Text	No	
Analytics	Bucket total	Total detail	spotify bulan ini total berapa	Category Detail Report	Text	No	
Analytics	Bucket count	Count detail	berapa transaksi spotify bulan ini	Category Detail Report	Text	No	
Analytics	Top transaction	Transaksi terbesar	entertainment terbesar bulan ini apa	Category Detail Report	Text	No	
Analytics	Average	Rata-rata bucket/detail	spotify rata-rata per bulan berapa	Category Detail Report	Text	No	
Analytics	Share	Kontribusi item	spotify kontribusinya berapa persen dari entertainment bulan ini	Category Detail Report	Text	No	
Analytics	Top merchant amount	Merchant terbesar	3 merchant entertainment terbesar bulan ini apa aja	Category Detail Report	Text	No	
Analytics	Top merchant frequency	Merchant paling sering	merchant entertainment paling sering bulan ini	Category Detail Report	Text	No	
Analytics	Explain change	Kenapa bucket naik	kenapa entertainment naik bulan ini	Category Detail Report	Text	No	
Analytics	Top category increase	Kategori paling naik	kategori mana yang paling naik dibanding bulan lalu	General Analytics	Text	No	
Analytics	Top merchant delta	Merchant pendorong kenaikan	merchant apa yang paling ngedorong kenaikan spending	General Analytics	Text	No	
Analytics	Recurring	Recurring report	top recurring expense bulan ini	General Analytics	Text	No	
Analytics	New merchants	Merchant baru	merchant baru bulan ini apa aja	General Analytics	Text	No	
Analytics	Weekend vs weekday	Banding weekend vs weekday	weekend lebih boros gak	General Analytics	Text	No	
Analytics	Habit leaks	Leak / bocor halus	kebiasaan bocor halus aku apa	General Analytics	Text	No	
Cashflow	Safety to payday	Aman sampai payday	aman sampai gajian gak	Cashflow Forecast	Text	No	
Cashflow	Safety to payday	Aman sampai payday	cukup nggak sampai payday	Cashflow Forecast	Text	No	
Cashflow	Month-end remaining	Sisa sampai akhir bulan	ujung bulan kira-kira sisa uang berapa	Cashflow Forecast	Text	No	
Cashflow	Scenario expense	Impact skenario expense	kalau bayar cicilan 1 juta besok masih aman gak	Cashflow Forecast	Text	No	
Goal	Set generic goal	Set target tabungan	mau nabung 50 juta	Goal	Text	No	
Goal	Set typed goal	Set goal spesifik	target rumah 800 juta	Goal	Text	No	
Goal	Status	Lihat status goal	status goal rumah gimana	Goal	Text	No	
Goal	Contribution	Tambah kontribusi	setor 500rb ke rumah	Goal	Text	No	
Goal Planner	Focus	Fokus goal	kalau fokus rumah dulu gimana	Goal Planner	Text	No	
Goal Planner	Focus duration	Fokus selama periode	kalau fokus rumah 6 bulan dulu gimana	Goal Planner	Text	No	
Goal Planner	Split	Bagi tabungan	kalau tabungan dibagi 60:40 hasilnya gimana	Goal Planner	Text	No	
Goal Planner	Priority	Target paling realistis	target mana yang paling realistis dulu	Goal Planner	Text	No	
Smart Allocation	Advice	Arahkan sisa uang	sisa uang bulan ini sebaiknya kemana	Smart Allocation	Text	No	Masih rule-based
Financial Freedom	Status	Tracker financial freedom	financial freedom aku aman gak	Financial Freedom	Text	No	Masih baseline planner
Wealth Projection	Future value	Proyeksi wealth	kalau invest 3 juta per bulan 10 tahun hasilnya berapa	Wealth Projection	Text	No	
Portfolio	Add asset	Catat aset	tambah saham BBCA 10 lot harga 9000	Portfolio	Text	No	
Portfolio	Summary	Ringkasan portfolio	portfolio aku gimana	Portfolio	Text	No	
Portfolio	Risk/rebalance	Analisa konsentrasi	perlu rebalance gak	Portfolio	Text	No	
Market	Price lookup	Harga market live	harga BBCA hari ini	Market	Text	Yes	Perlu provider API
Market	Price lookup	Harga market live	BTC sekarang berapa	Market	Text	Yes	Perlu provider API
News	Finance digest	Berita finance	berita finance pagi ini	News	Text	Yes	Perlu provider API
News	Portfolio news	News sesuai aset	ada news penting tentang aset aku gak?	News	Text	Yes	Perlu provider API
Reminder	Status	Lihat setting reminder	status reminder	Reminder Preference	Text	No	
Reminder	Disable type	Matikan tipe reminder	matikan reminder budget	Reminder Preference	Text	No	
Reminder	Cap per day	Batasi reminder	batasi reminder 2 per hari	Reminder Preference	Text	No	
Reminder	Pause	Pause sementara	pause reminder 12 jam	Reminder Preference	Text	No	
General Chat	Greeting	Sapaan natural	halo	General Chat	Text	No	
General Chat	Capabilities	Tanya kemampuan	kamu bisa apa?	General Chat	Text	No	
General Chat	Advice general	Tanya saran finansial	keuangan aku sehat gak?	Advice / Insight	Text	No	
General Chat	Safe fallback	Out of scope	cuaca hari ini gimana	General Chat Safe	Text	No	Bot harus jujur
Conversation Memory	Follow-up	Follow-up dari konteks	lanjut	Conversation Memory	Text	No	Harus ada konteks sebelumnya
Conversation Memory	Follow-up	Follow-up dari konteks	yang kedua	Conversation Memory	Text	No	Harus ada konteks sebelumnya
""", len(QUERY_HEADERS))

DATA_SHEETS = {
    '01_First_User_Journey': FIRST_USER,
    '02_Onboarding': ONBOARDING,
    '03_Transactions': TRANSACTIONS,
    '04_Reports_Analytics': REPORTS,
    '05_Planning_Goals': PLANNING,
    '06_Portfolio_Market_News': PORTFOLIO,
    '07_Reminders_General': REMINDERS,
}

base = Path(__file__).resolve().parent.parent
out_dir = base / 'artifacts'
out_dir.mkdir(exist_ok=True)
out_file = out_dir / 'ai-finance-assistant-testcases.xlsx'

wb = Workbook()
wb.remove(wb.active)
summary = wb.create_sheet('00_Summary')
summary['A1'] = 'AI Finance Assistant - Comprehensive Test Case Workbook'
summary['A1'].font = Font(size=16, bold=True)
summary['A3'] = 'Generated At'; summary['B3'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
summary['A4'] = 'Scope'; summary['B4'] = 'User first arrival, onboarding, transactions, reports, analytics, planning, goals, portfolio, reminders, general chat, and query catalog.'
summary['A5'] = 'Notes'; summary['B5'] = 'Rows bertanda Requires API dipakai saat provider market/news/gold sudah aktif.'
summary.append([]); summary.append(['Sheet','Purpose','Test Case Count'])
for cell in summary[7]:
    cell.fill = HEADER_FILL; cell.font = HEADER_FONT; cell.border = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
summary_rows = [
    ('01_First_User_Journey','End-to-end first user journey',len(FIRST_USER)),
    ('02_Onboarding','Branching onboarding and summary',len(ONBOARDING)),
    ('03_Transactions','Transaction, OCR, mutation, categorization',len(TRANSACTIONS)),
    ('04_Reports_Analytics','Report, analytics, cashflow, health score',len(REPORTS)),
    ('05_Planning_Goals','Budget, goals, allocation, FF, projection',len(PLANNING)),
    ('06_Portfolio_Market_News','Portfolio, market, news, fallback',len(PORTFOLIO)),
    ('07_Reminders_General','Reminders, general chat, memory safety',len(REMINDERS)),
    ('08_Query_Catalog','Query variants catalog',len(QUERY_ROWS)),
]
for item in summary_rows:
    summary.append(list(item))
set_widths(summary, {'A':28,'B':78,'C':18}); summary.freeze_panes = 'A7'

for name, rows in DATA_SHEETS.items():
    ws = wb.create_sheet(name)
    ws.append(COMMON_HEADERS)
    for row in rows:
        ws.append(row)
    style_sheet(ws, common=True)
    set_widths(ws, {'A':14,'B':18,'C':18,'D':24,'E':14,'F':10,'G':26,'H':34,'I':24,'J':36,'K':32,'L':24})

qs = wb.create_sheet('08_Query_Catalog')
qs.append(QUERY_HEADERS)
for row in QUERY_ROWS:
    qs.append(row)
style_sheet(qs, common=False, api_col=7)
set_widths(qs, {'A':18,'B':22,'C':26,'D':34,'E':22,'F':18,'G':12,'H':24})

for sheet in wb.worksheets:
    sheet.sheet_view.showGridLines = True
wb.save(out_file)
loaded = load_workbook(out_file)
print(out_file)
print(','.join(loaded.sheetnames))
for name in loaded.sheetnames:
    ws = loaded[name]
    print(f'{name}:{ws.max_row - 1}')
